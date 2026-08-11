"""Excel de desplacaments del C.B. Banyoles.

Regla del fitxer: cap import calculat no s'hi escriu com a numero. Tot el que es
deriva d'una altra cosa (km d'anada i tornada, tarifa aplicable, quilometratge,
dietes, totals, subtotals) hi va com a formula, de manera que es pugui auditar i
recalcular canviant els parametres del full "Parametres".

Dades primaries: data, equip, competicio, divisio, grup, club visitat i km d'anada
(mesura d'OSRM sobre OpenStreetMap).
"""

import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SP = Path(sys.argv[1])
# Amb l'argument "fiscal" el resum s'agrupa per any natural en comptes de per temporada.
FISCAL = len(sys.argv) > 2 and sys.argv[2] == "fiscal"
d = json.loads((SP / "rows.json").read_text(encoding="utf-8"))
ROWS = [r for r in d["rows"] if r["estat"] in ("verificat", "incompareixenca")]
PROJ, ADRECES, P = d["projeccio"], d["adreces"], d["params"]
INCOMP, CASA = d["incompar"], d["a_casa"]

VERD = "1F5F45"
HDR = PatternFill("solid", fgColor=VERD)
SUB = PatternFill("solid", fgColor="E6EFE9")
BOLD_W = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="C3CCC4")
BOX = Border(bottom=THIN)
EUR = '#,##0.00 "€"'
KM = '#,##0.0'
DATA = "DD/MM/YYYY"

wb = Workbook()


def capcalera(ws, noms, amples):
    ws.append(noms)
    for i, (n, a) in enumerate(zip(noms, amples), start=1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font = HDR, BOLD_W
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


# ---------------------------------------------------------------- Parametres
ws = wb.active
ws.title = "Paràmetres"
capcalera(ws, ["Paràmetre", "Valor", "Unitat", "Font"], [34, 14, 14, 62])
files = [
    ("Dieta de manutenció de migdia", P["dieta"], "€/jugador",
     "Import que aplica el club; per sota del límit exempt de 26,67 € de l'art. 9 del "
     "Reglament de l'IRPF"),
    ("Llindar de distància per meritar dieta", P["llindar_km"], "km d'anada",
     "Criteri del club"),
    ("Jugadors per desplaçament de lliga", P["jugadors_lliga"], "persones",
     "Dada del club; coincideix amb les actes de la federació"),
    ("Jugadors per desplaçament de Copa", P["jugadors_copa"], "persones",
     "Dada del club; a la Copa cada equip presenta 3 jugadors"),
    ("Barem de quilometratge fins al 16/07/2023", 0.19, "€/km",
     "Reglament de l'IRPF, redacció anterior a l'Ordre HFP/792/2023"),
    ("Barem de quilometratge des del 17/07/2023", 0.26, "€/km",
     "Ordre HFP/792/2023, de 12 de juliol (BOE 169, de 17/07/2023)"),
    ("Data d'entrada en vigor del barem nou", date(2023, 7, 17), "data",
     "Ordre HFP/792/2023"),
]
for f in files:
    ws.append(list(f))
for r in range(2, 2 + len(files)):
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
ws["B8"].number_format = DATA
DIETA, LLINDAR = "Paràmetres!$B$2", "Paràmetres!$B$3"
J_LLIGA, J_COPA = "Paràmetres!$B$4", "Paràmetres!$B$5"
T_VELLA, T_NOVA, D_CANVI = "Paràmetres!$B$6", "Paràmetres!$B$7", "Paràmetres!$B$8"
ws.append([])
ws.append(["Nota", "Els fulls de dades no contenen cap import escrit a mà: tot el que es "
                   "calcula hi va com a fórmula i es recalcula si es canvia un paràmetre "
                   "d'aquest full."])
ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True)

# ---------------------------------------------------------------- Clubs
wsc = wb.create_sheet("Clubs")
capcalera(wsc, ["Club", "Adreça", "Municipi", "km anada", "km anada i tornada",
                "Precisió de la coordenada"], [26, 46, 24, 11, 15, 20])
clubs = {}
for r in ROWS + PROJ:
    clubs.setdefault(r["seu_club"] if "seu_club" in r else r["rival_club"], r)
for i, (nom, r) in enumerate(sorted(clubs.items(), key=lambda x: x[1]["km_anada"]), start=2):
    wsc.append([nom, ADRECES.get(nom, ""), r["municipi"], r["km_anada"], None,
                r.get("precisio", "carrer")])
    wsc.cell(row=i, column=5).value = f"=D{i}*2"
    wsc.cell(row=i, column=4).number_format = KM
    wsc.cell(row=i, column=5).number_format = KM
N_CLUBS = wsc.max_row


def col_km_club(cell_club: str) -> str:
    """Busca els km d'anada del club a la taula Clubs, en comptes de repetir-los."""
    return f"VLOOKUP({cell_club},Clubs!$A$2:$D${N_CLUBS},4,FALSE)"


# ---------------------------------------------------------------- Desplacaments
wsd = wb.create_sheet("Desplaçaments")
COLS = ["Temporada", "Data", "Equip", "Competició", "Divisió", "Grup", "Club visitat",
        "Municipi", "km anada", "km anada i tornada", "Jugadors", "Tarifa (€/km)",
        "Quilometratge (€)", "Dieta (€)", "Total (€)", "Disputat", "Any"]
capcalera(wsd, COLS, [12, 11, 13, 24, 16, 14, 24, 22, 10, 15, 10, 12, 14, 11, 12, 10, 8])
for i, r in enumerate(ROWS, start=2):
    wsd.append([r["temporada"], date.fromisoformat(r["data"]), r["equip"],
                "Copa Catalana per equips" if r["tipus"] == "copa"
                else "Lliga Catalana Tres Bandes",
                r["divisio"], r["grup"], r["seu_club"], r["municipi"], None])
    wsd.cell(row=i, column=2).number_format = DATA
    wsd.cell(row=i, column=9).value = f"={col_km_club(f'G{i}')}"
    wsd.cell(row=i, column=16).value = "no" if r["estat"] == "incompareixenca" else "sí"
    wsd.cell(row=i, column=17).value = f"=YEAR(B{i})"
    wsd.cell(row=i, column=10).value = f"=I{i}*2"
    wsd.cell(row=i, column=11).value = (
        f'=IF(P{i}="no",0,IF(D{i}="Copa Catalana per equips",{J_COPA},{J_LLIGA}))')
    wsd.cell(row=i, column=12).value = f'=IF(P{i}="no",0,IF(B{i}>={D_CANVI},{T_NOVA},{T_VELLA}))'
    wsd.cell(row=i, column=13).value = f"=ROUND(J{i}*L{i},2)"
    wsd.cell(row=i, column=14).value = f"=IF(I{i}>{LLINDAR},ROUND(K{i}*{DIETA},2),0)"
    wsd.cell(row=i, column=15).value = f"=M{i}+N{i}"
    for c, fmt in ((9, KM), (10, KM), (12, EUR), (13, EUR), (14, EUR), (15, EUR)):
        wsd.cell(row=i, column=c).number_format = fmt
N_D = wsd.max_row
DISP = f'Desplaçaments!$P$2:$P${N_D},"sí"'
wsd.append([None] * 8 + [None, f'=SUMIFS(J2:J{N_D},{DISP})', None, None,
                         f"=SUM(M2:M{N_D})", f"=SUM(N2:N{N_D})", f"=SUM(O2:O{N_D})"])
tot = wsd.max_row
wsd.cell(row=tot, column=1).value = "TOTAL"
for c in range(1, 18):
    wsd.cell(row=tot, column=c).font = Font(bold=True)
    wsd.cell(row=tot, column=c).fill = SUB
for c, fmt in ((10, KM), (13, EUR), (14, EUR), (15, EUR)):
    wsd.cell(row=tot, column=c).number_format = fmt
wsd.auto_filter.ref = f"A1:Q{N_D}"

# ---------------------------------------------------------------- Resum
COL_PER, ETIQ = ("Q", "Any") if FISCAL else ("A", "Temporada")
wsr = wb.create_sheet(f"Resum per {'any fiscal' if FISCAL else 'temporada'}")
capcalera(wsr, [ETIQ, "Equips", "Desplaçaments", "km", "Quilometratge (€)",
                "Dietes (€)", "Total (€)"], [14, 16, 15, 12, 16, 13, 14])


def _per(r):
    return int(r["data"][:4]) if FISCAL else r["temporada"]


# La 2020-2021 no te desplacaments pero es llista igualment, al seu lloc.
temps = (sorted({_per(r) for r in ROWS}) if FISCAL
         else sorted({r["temporada"] for r in ROWS} | {"2020-2021"}))
for i, t in enumerate(temps, start=2):
    eq = sorted({r["equip"].replace("Banyoles ", "").replace('"', "")
                 for r in ROWS if _per(r) == t and r["tipus"] == "regular"})
    wsr.append([t, ", ".join(eq) if eq else "cap"])
    rng = f"Desplaçaments!${COL_PER}$2:${COL_PER}${N_D}"
    wsr.cell(row=i, column=3).value = (
        f'=COUNTIFS({rng},A{i},Desplaçaments!$P$2:$P${N_D},"sí")')
    wsr.cell(row=i, column=4).value = (
        f'=SUMIFS(Desplaçaments!$J$2:$J${N_D},{rng},A{i},'
        f'Desplaçaments!$P$2:$P${N_D},"sí")')
    for col, src in ((5, "M"), (6, "N"), (7, "O")):
        wsr.cell(row=i, column=col).value = (
            f'=SUMIF({rng},A{i},Desplaçaments!${src}$2:${src}${N_D})')
    for c, fmt in ((4, KM), (5, EUR), (6, EUR), (7, EUR)):
        wsr.cell(row=i, column=c).number_format = fmt
n = wsr.max_row
wsr.append(["TOTAL", None, f"=SUM(C2:C{n})", f"=SUM(D2:D{n})", f"=SUM(E2:E{n})",
            f"=SUM(F2:F{n})", f"=SUM(G2:G{n})"])
for c in range(1, 8):
    wsr.cell(row=wsr.max_row, column=c).font = Font(bold=True)
    wsr.cell(row=wsr.max_row, column=c).fill = SUB
for c, fmt in ((4, KM), (5, EUR), (6, EUR), (7, EUR)):
    wsr.cell(row=wsr.max_row, column=c).number_format = fmt

# ---------------------------------------------------------------- Projeccio
wsp = wb.create_sheet("Projecció 2026-2027")
capcalera(wsp, ["Equip", "Divisió", "Grup", "Club rival", "Equip rival", "Municipi",
                "km anada", "km anada i tornada", "Jugadors", "Tarifa (€/km)",
                "Quilometratge (€)", "Dieta (€)", "Total (€)"],
          [13, 14, 8, 24, 12, 22, 10, 15, 10, 12, 15, 11, 12])
for i, r in enumerate(PROJ, start=2):
    wsp.append([r["equip"], r["divisio"], r["grup"], r["rival_club"], r["rival_equip"],
                r["municipi"]])
    wsp.cell(row=i, column=7).value = f"={col_km_club(f'D{i}')}"
    wsp.cell(row=i, column=8).value = f"=G{i}*2"
    wsp.cell(row=i, column=9).value = f"={J_LLIGA}"
    wsp.cell(row=i, column=10).value = f"={T_NOVA}"
    wsp.cell(row=i, column=11).value = f"=ROUND(H{i}*J{i},2)"
    wsp.cell(row=i, column=12).value = f"=IF(G{i}>{LLINDAR},ROUND(I{i}*{DIETA},2),0)"
    wsp.cell(row=i, column=13).value = f"=K{i}+L{i}"
    for c, fmt in ((7, KM), (8, KM), (10, EUR), (11, EUR), (12, EUR), (13, EUR)):
        wsp.cell(row=i, column=c).number_format = fmt
n = wsp.max_row
wsp.append(["TOTAL"] + [None] * 6 + [f"=SUM(H2:H{n})", None, None,
                                     f"=SUM(K2:K{n})", f"=SUM(L2:L{n})", f"=SUM(M2:M{n})"])
for c in range(1, 14):
    wsp.cell(row=wsp.max_row, column=c).font = Font(bold=True)
    wsp.cell(row=wsp.max_row, column=c).fill = SUB
for c, fmt in ((8, KM), (11, EUR), (12, EUR), (13, EUR)):
    wsp.cell(row=wsp.max_row, column=c).number_format = fmt

# ---------------------------------------------------------------- Notes
wsn = wb.create_sheet("Notes")
capcalera(wsn, ["Concepte", "Detall"], [30, 105])
notes = [
    ("Àmbit", "Competicions oficials per equips de la Federació Catalana de Billar: "
              "Lliga Catalana Tres Bandes, incloses fases finals i promocions, i Copa "
              "Catalana per equips. No s'hi inclouen les proves individuals."),
    ("Període", "Anys naturals 2014 a 2026. La temporada esportiva va de setembre a maig, "
                "de manera que cada any natural recull el tram final d'una temporada i el "
                "començament de la següent." if FISCAL else
                "Temporades 2014-2015 a 2025-2026. El club no va tenir equips en "
                "competició la 2020-2021."),
    ("Encontres no disputats", "Els que porten «no» a la columna Disputat consten al "
                               "calendari federatiu amb resultat d'incompareixença: no es "
                               "van jugar i el full els deixa a zero."),
    ("Unitat", "Un desplaçament és un viatge d'anada i tornada per a una jornada, comptat "
               "només quan es juga fora de casa."),
    ("Calendari i resultats", "Portal públic de la Federació Catalana de Billar "
                              "(www.fcbillar.cat), verificat grup a grup i jornada a jornada. De la "
                              "2014-2015 a la 2019-2020 les dades venen directament "
                              "del portal, perquè la base de dades del club no les té "
                              "completes."),
    ("Adreces", "Directori oficial de clubs de la FCB, unificades a un sol criteri "
                "tipogràfic sense alterar-ne el contingut."),
    ("Distàncies", "OSRM sobre la xarxa viària d'OpenStreetMap, perfil de vehicle, ruta més "
                   "ràpida, des del Club Billar Banyoles (carrer de l'Abeurador, 10)."),
    ("Precisió", "28 clubs estan situats a nivell de carrer o portal. Canet de Mar, Sant "
                 "Feliu de Codines i Cardona, a nivell de municipi, perquè el carrer indicat "
                 "no consta a la cartografia."),
    ("Quilometratge", "Barem exempt de gravamen a l'IRPF vigent el dia del partit. No s'hi "
                      "han afegit peatges ni aparcament."),
    ("Projecció 2026-2027", "Previsió, no despesa realitzada. Composició de grups derivada "
                            "de les classificacions oficials 2025-2026 i dels play-offs de "
                            "promoció del 4 i 5 de juliol de 2026. No inclou la Copa."),
]
for c, t in notes:
    wsn.append([c, t])
for r in range(2, wsn.max_row + 1):
    wsn.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    wsn.cell(row=r, column=1).font = Font(bold=True)
    wsn.row_dimensions[r].height = 30

wsn.append([])
for r in CASA:
    wsn.append([r["data_dmy"], f'{r["equip"]} · {r["divisio"]} {r["grup"]} · '
                               f'jugat a Banyoles: {r["nota"]}'])

for sheet in wb:
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for c in row:
            if c.row > 1:
                c.border = BOX

out = SP / ("Desplacaments_CB_Banyoles_any_fiscal.xlsx" if FISCAL
            else "Desplacaments_CB_Banyoles.xlsx")
wb.save(out)
print(f"{out.name}: {len(ROWS)} desplaçaments, {len(PROJ)} de projecció, "
      f"{N_CLUBS - 1} clubs, {len(wb.sheetnames)} fulls")
